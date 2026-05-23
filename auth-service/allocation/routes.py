import io
import json
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from auth.dependencies import get_current_user, require_manager_or_above, require_sme_or_above
from business_requirements.models import BRStream, BusinessRequirement
from database import get_db
from models import Role, User
from notifications.models import NotificationType
from notifications.service import create_notification
from streams.models import BatchStream, BatchStreamSME

from .ai_recommender import generate_allocation_recommendations
from .models import AllocationAIRecommendation, AllocationConfig, RequestStatus, SMEAssociateRequest, TraineeAllocation
from .schemas import (
    AllocationAIRecommendationResponse,
    AllocationConfigResponse,
    AllocationConfigUpdate,
    AllocationRunRequest,
    AllocationRunResult,
    ManualOverrideRequest,
    SMEAssociateRequestCreate,
    SMEAssociateRequestResponse,
    SMEAssociateRequestReview,
    StreamScoreDetail,
    TraineeAllocationResponse,
)
from .service import get_or_create_config, run_allocation

router = APIRouter(prefix="/allocation", tags=["allocation"])


def _stream_name(stream_id: int | None, db: Session) -> str | None:
    if stream_id is None:
        return None
    s = db.query(BatchStream).filter_by(id=stream_id).first()
    return s.name if s else None


def _build_response(alloc: TraineeAllocation, db: Session) -> TraineeAllocationResponse:
    breakdown: dict[str, float] = {}
    try:
        breakdown = json.loads(alloc.score_breakdown_json or "{}")
    except (ValueError, TypeError):
        pass

    stream_scores_raw: dict[str, dict] = {}
    try:
        stream_scores_raw = json.loads(alloc.all_stream_scores_json or "{}")
    except (ValueError, TypeError):
        pass

    all_stream_scores: list[StreamScoreDetail] = []
    for sid_str, vals in stream_scores_raw.items():
        try:
            sid = int(sid_str)
        except ValueError:
            continue
        name = _stream_name(sid, db) or sid_str
        all_stream_scores.append(
            StreamScoreDetail(
                stream_id=sid,
                stream_name=name,
                composite=vals.get("composite", 0.0),
                subject_score=vals.get("subject_score", 0.0),
            )
        )
    all_stream_scores.sort(key=lambda x: x.composite, reverse=True)

    effective_id = alloc.manual_stream_id if alloc.manual_stream_id is not None else alloc.suggested_stream_id

    return TraineeAllocationResponse(
        id=alloc.id,
        batch_name=alloc.batch_name,
        employee_id=alloc.employee_id,
        trainee_name=alloc.trainee_name,
        dpi_score=alloc.dpi_score,
        subject_score=alloc.subject_score,
        composite_score=alloc.composite_score,
        suggested_stream_id=alloc.suggested_stream_id,
        suggested_stream_name=_stream_name(alloc.suggested_stream_id, db),
        manual_stream_id=alloc.manual_stream_id,
        manual_stream_name=_stream_name(alloc.manual_stream_id, db),
        effective_stream_id=effective_id,
        effective_stream_name=_stream_name(effective_id, db),
        manual_override_reason=alloc.manual_override_reason,
        overridden_by_email=alloc.overridden_by_email,
        overridden_at=alloc.overridden_at,
        is_frozen=alloc.is_frozen,
        frozen_at=alloc.frozen_at,
        frozen_by_email=alloc.frozen_by_email,
        score_breakdown=breakdown,
        all_stream_scores=all_stream_scores,
    )


# ── Config ──────────────────────────────────────────────────────────────────

@router.get("/{batch_name}/config", response_model=AllocationConfigResponse)
def get_config(
    batch_name: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    return get_or_create_config(batch_name, db)


@router.put("/{batch_name}/config", response_model=AllocationConfigResponse)
def update_config(
    batch_name: str,
    body: AllocationConfigUpdate,
    db: Session = Depends(get_db),
    _=Depends(require_manager_or_above),
):
    cfg = get_or_create_config(batch_name, db)
    if cfg.is_frozen:
        raise HTTPException(status_code=status.HTTP_423_LOCKED, detail="Allocation is frozen. Unfreeze the batch before changing weights.")
    cfg.score_weight = body.score_weight
    cfg.dpi_weight = body.dpi_weight
    db.commit()
    db.refresh(cfg)
    return cfg


# ── Run ─────────────────────────────────────────────────────────────────────

@router.post("/{batch_name}/run", response_model=AllocationRunResult)
def trigger_allocation(
    batch_name: str,
    body: AllocationRunRequest = AllocationRunRequest(),
    db: Session = Depends(get_db),
    current_user=Depends(require_manager_or_above),
):
    cfg = get_or_create_config(batch_name, db)
    if cfg.is_frozen:
        raise HTTPException(status_code=status.HTTP_423_LOCKED, detail="Allocation is frozen. Unfreeze the batch before re-running.")
    result = run_allocation(batch_name, current_user.email, db, mode=body.mode)
    return result


# ── Results ─────────────────────────────────────────────────────────────────

@router.get("/{batch_name}", response_model=list[TraineeAllocationResponse])
def get_allocations(
    batch_name: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    rows = (
        db.query(TraineeAllocation)
        .filter_by(batch_name=batch_name)
        .order_by(TraineeAllocation.trainee_name)
        .all()
    )
    return [_build_response(r, db) for r in rows]


# ── Manual override ──────────────────────────────────────────────────────────

@router.patch("/{batch_name}/{employee_id}/override", response_model=TraineeAllocationResponse)
def set_override(
    batch_name: str,
    employee_id: str,
    body: ManualOverrideRequest,
    db: Session = Depends(get_db),
    current_user=Depends(require_manager_or_above),
):
    alloc = db.query(TraineeAllocation).filter_by(
        batch_name=batch_name, employee_id=employee_id
    ).first()
    if not alloc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Trainee allocation not found")

    cfg = get_or_create_config(batch_name, db)
    if cfg.is_frozen:
        raise HTTPException(status_code=status.HTTP_423_LOCKED, detail="Batch allocation is frozen. Unfreeze before making overrides.")
    if alloc.is_frozen:
        raise HTTPException(status_code=status.HTTP_423_LOCKED, detail="This trainee's allocation is frozen. Unfreeze before overriding.")

    stream = db.query(BatchStream).filter_by(id=body.stream_id, batch_name=batch_name).first()
    if not stream:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Stream not found in this batch")

    alloc.manual_stream_id = body.stream_id
    alloc.manual_override_reason = body.reason
    alloc.overridden_by_email = current_user.email
    alloc.overridden_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(alloc)
    return _build_response(alloc, db)


@router.delete("/{batch_name}/{employee_id}/override", response_model=TraineeAllocationResponse)
def clear_override(
    batch_name: str,
    employee_id: str,
    db: Session = Depends(get_db),
    _=Depends(require_manager_or_above),
):
    alloc = db.query(TraineeAllocation).filter_by(
        batch_name=batch_name, employee_id=employee_id
    ).first()
    if not alloc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Trainee allocation not found")

    cfg = get_or_create_config(batch_name, db)
    if cfg.is_frozen:
        raise HTTPException(status_code=status.HTTP_423_LOCKED, detail="Batch allocation is frozen. Unfreeze before clearing overrides.")
    if alloc.is_frozen:
        raise HTTPException(status_code=status.HTTP_423_LOCKED, detail="This trainee's allocation is frozen. Unfreeze before clearing the override.")

    alloc.manual_stream_id = None
    alloc.manual_override_reason = None
    alloc.overridden_by_email = None
    alloc.overridden_at = None
    db.commit()
    db.refresh(alloc)
    return _build_response(alloc, db)


# ── Freeze / Unfreeze ────────────────────────────────────────────────────────

@router.post("/{batch_name}/freeze", response_model=AllocationConfigResponse)
def freeze_batch(
    batch_name: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_manager_or_above),
):
    cfg = get_or_create_config(batch_name, db)
    if cfg.is_frozen:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Batch allocation is already frozen.")
    cfg.is_frozen = True
    cfg.frozen_at = datetime.now(timezone.utc)
    cfg.frozen_by_email = current_user.email
    db.commit()
    db.refresh(cfg)
    return cfg


@router.post("/{batch_name}/unfreeze", response_model=AllocationConfigResponse)
def unfreeze_batch(
    batch_name: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_manager_or_above),
):
    cfg = get_or_create_config(batch_name, db)
    if not cfg.is_frozen:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Batch allocation is not frozen.")
    cfg.is_frozen = False
    cfg.frozen_at = None
    cfg.frozen_by_email = None
    db.commit()
    db.refresh(cfg)
    return cfg


@router.post("/{batch_name}/{employee_id}/freeze", response_model=TraineeAllocationResponse)
def freeze_trainee(
    batch_name: str,
    employee_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_manager_or_above),
):
    alloc = db.query(TraineeAllocation).filter_by(batch_name=batch_name, employee_id=employee_id).first()
    if not alloc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Trainee allocation not found.")
    if alloc.is_frozen:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Trainee allocation is already frozen.")
    alloc.is_frozen = True
    alloc.frozen_at = datetime.now(timezone.utc)
    alloc.frozen_by_email = current_user.email
    db.commit()
    db.refresh(alloc)
    return _build_response(alloc, db)


@router.post("/{batch_name}/{employee_id}/unfreeze", response_model=TraineeAllocationResponse)
def unfreeze_trainee(
    batch_name: str,
    employee_id: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_manager_or_above),
):
    alloc = db.query(TraineeAllocation).filter_by(batch_name=batch_name, employee_id=employee_id).first()
    if not alloc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Trainee allocation not found.")
    if not alloc.is_frozen:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Trainee allocation is not frozen.")
    alloc.is_frozen = False
    alloc.frozen_at = None
    alloc.frozen_by_email = None
    db.commit()
    db.refresh(alloc)
    return _build_response(alloc, db)


# ── Excel Export ─────────────────────────────────────────────────────────────

@router.get("/{batch_name}/export")
def export_allocation(
    batch_name: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    rows = (
        db.query(TraineeAllocation)
        .filter_by(batch_name=batch_name)
        .order_by(TraineeAllocation.trainee_name)
        .all()
    )
    if not rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No allocations found for this batch.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Allocation"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    center = Alignment(horizontal="center", vertical="center")
    frozen_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")

    headers = [
        "Employee ID", "Trainee Name", "Batch", "DPI Score",
        "Subject Score", "Composite Score",
        "Suggested Stream", "Manual Override Stream",
        "Effective Stream", "Override Reason", "Overridden By",
        "Frozen", "Frozen By",
    ]

    # Collect all subject names for dynamic columns
    all_subjects: set[str] = set()
    for row in rows:
        try:
            breakdown = json.loads(row.score_breakdown_json or "{}")
            all_subjects.update(breakdown.keys())
        except (ValueError, TypeError):
            pass
    sorted_subjects = sorted(all_subjects)
    full_headers = headers + [f"Score: {s.title()}" for s in sorted_subjects]

    for col_idx, h in enumerate(full_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, alloc in enumerate(rows, start=2):
        effective_id = alloc.manual_stream_id if alloc.manual_stream_id is not None else alloc.suggested_stream_id
        suggested_name = _stream_name(alloc.suggested_stream_id, db)
        manual_name = _stream_name(alloc.manual_stream_id, db)
        effective_name = _stream_name(effective_id, db)

        breakdown: dict[str, float] = {}
        try:
            breakdown = json.loads(alloc.score_breakdown_json or "{}")
        except (ValueError, TypeError):
            pass

        base_data = [
            alloc.employee_id,
            alloc.trainee_name,
            alloc.batch_name,
            alloc.dpi_score,
            round(alloc.subject_score, 2) if alloc.subject_score is not None else None,
            round(alloc.composite_score, 2) if alloc.composite_score is not None else None,
            suggested_name or "Unallocated",
            manual_name or "—",
            effective_name or "Unallocated",
            alloc.manual_override_reason or "—",
            alloc.overridden_by_email or "—",
            "Yes" if alloc.is_frozen else "No",
            alloc.frozen_by_email or "—",
        ]
        subject_data = [round(breakdown.get(s, 0), 2) for s in sorted_subjects]
        full_row = base_data + subject_data

        for col_idx, val in enumerate(full_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if alloc.is_frozen:
                cell.fill = frozen_fill

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = batch_name.replace(" ", "_").replace("/", "-")
    filename = f"allocation_{safe_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── AI Recommendations ───────────────────────────────────────────────────────

@router.post(
    "/{batch_name}/ai-recommendations/generate",
    response_model=list[AllocationAIRecommendationResponse],
    status_code=status.HTTP_201_CREATED,
)
async def generate_ai_recommendations(
    batch_name: str,
    db: Session = Depends(get_db),
    current_user=Depends(require_manager_or_above),
):
    allocs = (
        db.query(TraineeAllocation)
        .filter_by(batch_name=batch_name)
        .order_by(TraineeAllocation.trainee_name)
        .all()
    )
    if not allocs:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No allocations found for this batch. Run allocation first.",
        )

    streams = (
        db.query(BatchStream)
        .filter_by(batch_name=batch_name, is_active=True)
        .all()
    )
    stream_names = [s.name for s in streams]
    stream_name_to_id = {s.name: s.id for s in streams}

    brs = (
        db.query(BusinessRequirement)
        .filter(BusinessRequirement.batch_name == batch_name, BusinessRequirement.is_active == True)
        .all()
    )
    br_data = []
    for br in brs:
        br_streams = (
            db.query(BRStream)
            .filter(BRStream.br_id == br.id, BRStream.is_active == True)
            .all()
        )
        br_data.append({
            "title": br.title,
            "location": br.location,
            "streams": [
                {
                    "name": s.name,
                    "is_mandatory": s.is_mandatory,
                    "roles_needed": s.roles_needed,
                    "subjects_needed": s.subjects_needed,
                }
                for s in br_streams
            ],
        })

    trainee_inputs = []
    for alloc in allocs:
        score_breakdown = {}
        try:
            score_breakdown = json.loads(alloc.score_breakdown_json or "{}")
        except (ValueError, TypeError):
            pass

        stream_fit = []
        try:
            raw = json.loads(alloc.all_stream_scores_json or "{}")
            for sid_str, vals in raw.items():
                try:
                    sid = int(sid_str)
                except ValueError:
                    continue
                name = _stream_name(sid, db) or sid_str
                stream_fit.append({"stream": name, "composite": round(vals.get("composite", 0.0), 2)})
            stream_fit.sort(key=lambda x: x["composite"], reverse=True)
        except (ValueError, TypeError):
            pass

        trainee_inputs.append({
            "employee_id": alloc.employee_id,
            "name": alloc.trainee_name,
            "dpi_score": alloc.dpi_score,
            "algorithm_suggested_stream": _stream_name(alloc.suggested_stream_id, db),
            "subject_scores": score_breakdown,
            "stream_fit_scores": stream_fit,
        })

    try:
        recommendations = await generate_allocation_recommendations(
            batch_name=batch_name,
            trainees=trainee_inputs,
            streams=stream_names,
            business_requirements=br_data,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"AI recommendation failed — {exc}",
        )

    # Replace existing recommendations for this batch
    db.query(AllocationAIRecommendation).filter_by(batch_name=batch_name).delete()

    gen_id = str(uuid.uuid4())
    rec_by_eid = {r["employee_id"]: r for r in recommendations}

    rows: list[AllocationAIRecommendation] = []
    for alloc in allocs:
        rec = rec_by_eid.get(alloc.employee_id)
        if not rec:
            continue
        row = AllocationAIRecommendation(
            batch_name=batch_name,
            employee_id=alloc.employee_id,
            generation_id=gen_id,
            agrees_with_algorithm=rec["agrees"],
            recommended_stream_name=rec.get("recommended_stream"),
            confidence=rec["confidence"],
            reasoning=rec["reasoning"],
            generated_by_email=current_user.email,
        )
        db.add(row)
        rows.append(row)

    db.commit()
    for row in rows:
        db.refresh(row)

    return rows


@router.get("/{batch_name}/ai-recommendations", response_model=list[AllocationAIRecommendationResponse])
def get_ai_recommendations(
    batch_name: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    return (
        db.query(AllocationAIRecommendation)
        .filter_by(batch_name=batch_name)
        .order_by(AllocationAIRecommendation.employee_id)
        .all()
    )


# ── SME Associate Requests ───────────────────────────────────────────────────

def _build_sme_request_response(req: SMEAssociateRequest, db: Session) -> SMEAssociateRequestResponse:
    try:
        requested = json.loads(req.requested_employee_ids or "[]")
    except (ValueError, TypeError):
        requested = []

    approved = None
    if req.approved_employee_ids is not None:
        try:
            approved = json.loads(req.approved_employee_ids)
        except (ValueError, TypeError):
            approved = []

    return SMEAssociateRequestResponse(
        id=req.id,
        batch_name=req.batch_name,
        stream_id=req.stream_id,
        stream_name=_stream_name(req.stream_id, db),
        sme_email=req.sme_email,
        requested_employee_ids=requested,
        status=req.status,
        approved_employee_ids=approved,
        reviewed_by_email=req.reviewed_by_email,
        reviewed_at=req.reviewed_at,
        review_notes=req.review_notes,
        created_at=req.created_at,
    )


@router.post(
    "/{batch_name}/sme-requests",
    response_model=SMEAssociateRequestResponse,
    status_code=status.HTTP_201_CREATED,
)
def create_sme_associate_request(
    batch_name: str,
    body: SMEAssociateRequestCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sme_or_above),
):
    stream = db.query(BatchStream).filter_by(id=body.stream_id, batch_name=batch_name).first()
    if not stream:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Stream not found in this batch")

    if current_user.role == Role.sme:
        assignment = db.query(BatchStreamSME).filter_by(
            stream_id=body.stream_id, user_id=current_user.id, is_active=True
        ).first()
        if not assignment:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You are not assigned to this stream",
            )

    existing_eids = {
        row.employee_id
        for row in db.query(TraineeAllocation.employee_id).filter_by(batch_name=batch_name).all()
    }
    invalid = [eid for eid in body.requested_employee_ids if eid not in existing_eids]
    if invalid:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Employee IDs not found in this batch: {invalid}",
        )

    req = SMEAssociateRequest(
        batch_name=batch_name,
        stream_id=body.stream_id,
        sme_user_id=current_user.id,
        sme_email=current_user.email,
        requested_employee_ids=json.dumps(body.requested_employee_ids),
        status=RequestStatus.pending,
    )
    db.add(req)
    db.flush()

    managers = (
        db.query(User)
        .filter(User.role.in_([Role.admin, Role.manager]), User.is_active == True)  # noqa: E712
        .all()
    )
    for mgr in managers:
        create_notification(
            db,
            mgr.email,
            NotificationType.sme_request_submitted,
            "SME Associate Request",
            f"{current_user.email} requested {len(body.requested_employee_ids)} associate(s) "
            f"for stream '{stream.name}' in batch '{batch_name}'.",
        )

    db.commit()
    db.refresh(req)
    return _build_sme_request_response(req, db)


@router.get("/{batch_name}/sme-requests", response_model=list[SMEAssociateRequestResponse])
def list_sme_requests(
    batch_name: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(SMEAssociateRequest).filter_by(batch_name=batch_name)
    if current_user.role == Role.sme:
        q = q.filter_by(sme_user_id=current_user.id)
    requests = q.order_by(SMEAssociateRequest.created_at.desc()).all()
    return [_build_sme_request_response(r, db) for r in requests]


@router.post("/{batch_name}/sme-requests/{request_id}/review", response_model=SMEAssociateRequestResponse)
def review_sme_request(
    batch_name: str,
    request_id: int,
    body: SMEAssociateRequestReview,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_or_above),
):
    req = db.query(SMEAssociateRequest).filter_by(id=request_id, batch_name=batch_name).first()
    if not req:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Request not found")
    if req.status != RequestStatus.pending:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Request has already been reviewed")

    requested = json.loads(req.requested_employee_ids)
    invalid = [eid for eid in body.approved_employee_ids if eid not in requested]
    if invalid:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Approved IDs not in original request: {invalid}",
        )

    if not body.approved_employee_ids:
        new_status = RequestStatus.rejected
    elif set(body.approved_employee_ids) == set(requested):
        new_status = RequestStatus.approved
    else:
        new_status = RequestStatus.partially_approved

    req.status = new_status
    req.approved_employee_ids = json.dumps(body.approved_employee_ids)
    req.reviewed_by_email = current_user.email
    req.reviewed_at = datetime.now(timezone.utc)
    req.review_notes = body.review_notes

    status_label = new_status.value.replace("_", " ").title()
    create_notification(
        db,
        req.sme_email,
        NotificationType.sme_request_reviewed,
        f"Associate Request {status_label}",
        f"Your request for {len(requested)} associate(s) in batch '{batch_name}' "
        f"was {status_label.lower()} by {current_user.email}.",
    )

    db.commit()
    db.refresh(req)
    return _build_sme_request_response(req, db)


@router.delete("/{batch_name}/sme-requests/{request_id}", status_code=status.HTTP_204_NO_CONTENT)
def cancel_sme_request(
    batch_name: str,
    request_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_sme_or_above),
):
    req = db.query(SMEAssociateRequest).filter_by(id=request_id, batch_name=batch_name).first()
    if not req:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Request not found")
    if current_user.role == Role.sme and req.sme_user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not your request")
    if req.status != RequestStatus.pending:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Only pending requests can be cancelled")

    req.status = RequestStatus.cancelled
    db.commit()
