import io
import json
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from auth.dependencies import get_current_user, require_manager_or_above
from database import get_db
from sync.models import SyncedBatch

from .models import BRStream, BusinessRequirement
from .schemas import (
    BRCreate,
    BRResponse,
    BRStreamCreate,
    BRStreamResponse,
    BRSummary,
    BRUpdate,
    ExcelImportResult,
)

router = APIRouter(prefix="/business-requirements", tags=["business-requirements"])


# ── Helpers ───────────────────────────────────────────────────────────

def _parse_list(raw: str) -> list[str]:
    try:
        return json.loads(raw)
    except Exception:
        return []


def _stream_resp(s: BRStream) -> BRStreamResponse:
    return BRStreamResponse(
        id=s.id,
        br_id=s.br_id,
        name=s.name,
        is_mandatory=s.is_mandatory,
        capacity_type=s.capacity_type,
        capacity_value=s.capacity_value,
        roles_needed=_parse_list(s.roles_needed),
        subjects_needed=_parse_list(s.subjects_needed),
        is_active=s.is_active,
    )


def _br_full(br: BusinessRequirement, db: Session) -> BRResponse:
    streams = (
        db.query(BRStream)
        .filter(BRStream.br_id == br.id, BRStream.is_active == True)
        .all()
    )
    return BRResponse(
        id=br.id,
        batch_name=br.batch_name,
        title=br.title,
        location=br.location,
        created_at=br.created_at,
        is_active=br.is_active,
        streams=[_stream_resp(s) for s in streams],
    )


def _bulk_insert_streams(br_id: int, stream_list: list[BRStreamCreate], db: Session) -> None:
    for s in stream_list:
        db.add(BRStream(
            br_id=br_id,
            name=s.name,
            is_mandatory=s.is_mandatory,
            capacity_type=s.capacity_type,
            capacity_value=s.capacity_value,
            roles_needed=json.dumps(s.roles_needed),
            subjects_needed=json.dumps(s.subjects_needed),
        ))


# ── Static routes — must be defined BEFORE /{br_id} ──────────────────

@router.get("/excel-template")
def download_template(_=Depends(get_current_user)):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Streams"

    headers = ["Stream Name", "Is Mandatory", "Capacity Type", "Capacity Value", "Roles Needed", "Subjects Needed"]
    hdr_fill = PatternFill(start_color="1E4D8C", end_color="1E4D8C", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    sample_rows = [
        ["Java Development",        "YES", "PERCENTAGE", 25, "Java Developer, Backend Engineer", "Java, SQL"],
        ["Python/Data Engineering",  "YES", "PERCENTAGE", 20, "Data Analyst, Data Engineer",      "Python"],
        ["Cloud & DevOps",           "YES", "PERCENTAGE", 20, "Cloud Engineer, DevOps Engineer",  "Cloud"],
        ["Cybersecurity",            "YES", "PERCENTAGE", 15, "Security Analyst",                 "Cybersecurity"],
        ["Business Analysis",        "YES", "PERCENTAGE", 10, "Business Analyst",                 "Agile"],
        ["AI/ML Engineering",        "YES", "PERCENTAGE", 10, "ML Engineer, AI Researcher",       "AIML, Python"],
    ]
    for row_data in sample_rows:
        ws.append(row_data)

    col_widths = [26, 15, 18, 16, 40, 30]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=br_template.xlsx"},
    )


@router.post("/parse-excel", response_model=ExcelImportResult)
async def parse_excel(file: UploadFile = File(...), _=Depends(require_manager_or_above)):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or corrupt Excel file")

    ws = wb.active
    parsed: list[BRStreamCreate] = []
    errors: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None for cell in row):
            continue

        def _get(i: int):
            return row[i] if len(row) > i else None

        name_val      = _get(0)
        mandatory_val = _get(1)
        cap_type_val  = _get(2)
        cap_value_val = _get(3)
        roles_val     = _get(4)
        subjects_val  = _get(5)

        if not name_val:
            errors.append(f"Row {row_idx}: Stream Name is required — skipped")
            continue

        is_mandatory = str(mandatory_val).strip().upper() == "YES" if mandatory_val else False

        cap_type_str = str(cap_type_val).strip().upper() if cap_type_val else ""
        if cap_type_str not in ("PERCENTAGE", "COUNT"):
            errors.append(f"Row {row_idx} ({name_val}): Capacity Type must be PERCENTAGE or COUNT")
            continue

        cap_type = "percentage" if cap_type_str == "PERCENTAGE" else "count"

        try:
            cap_value = float(cap_value_val)
        except (TypeError, ValueError):
            errors.append(f"Row {row_idx} ({name_val}): Capacity Value must be a number")
            continue

        roles    = [r.strip() for r in str(roles_val).split(",")    if r.strip()] if roles_val    else []
        subjects = [s.strip() for s in str(subjects_val).split(",") if s.strip()] if subjects_val else []

        parsed.append(BRStreamCreate(
            name=str(name_val).strip(),
            is_mandatory=is_mandatory,
            capacity_type=cap_type,
            capacity_value=cap_value,
            roles_needed=roles,
            subjects_needed=subjects,
        ))

    return ExcelImportResult(streams=parsed, errors=errors)


# ── List + Create ─────────────────────────────────────────────────────

@router.get("", response_model=list[BRSummary])
def list_brs(
    batch_name: str | None = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(BusinessRequirement).filter(BusinessRequirement.is_active == True)
    if batch_name:
        q = q.filter(BusinessRequirement.batch_name == batch_name)
    brs = q.order_by(BusinessRequirement.created_at.desc()).all()

    result = []
    for br in brs:
        count = (
            db.query(BRStream)
            .filter(BRStream.br_id == br.id, BRStream.is_active == True)
            .count()
        )
        result.append(BRSummary(
            id=br.id,
            batch_name=br.batch_name,
            title=br.title,
            location=br.location,
            created_at=br.created_at,
            is_active=br.is_active,
            stream_count=count,
        ))
    return result


@router.post("", response_model=BRResponse, status_code=status.HTTP_201_CREATED)
def create_br(body: BRCreate, db: Session = Depends(get_db), _=Depends(require_manager_or_above)):
    batch = db.query(SyncedBatch).filter(SyncedBatch.batch_name == body.batch_name).first()
    if not batch:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Batch '{body.batch_name}' not found in synced data",
        )
    br = BusinessRequirement(
        batch_name=body.batch_name,
        title=body.title,
        location=body.location,
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    db.add(br)
    db.flush()
    _bulk_insert_streams(br.id, body.streams, db)
    db.commit()
    db.refresh(br)
    return _br_full(br, db)


# ── Single BR CRUD ────────────────────────────────────────────────────

@router.get("/{br_id}", response_model=BRResponse)
def get_br(br_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    br = (
        db.query(BusinessRequirement)
        .filter(BusinessRequirement.id == br_id, BusinessRequirement.is_active == True)
        .first()
    )
    if not br:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Business requirement not found")
    return _br_full(br, db)


@router.put("/{br_id}", response_model=BRResponse)
def update_br(br_id: int, body: BRUpdate, db: Session = Depends(get_db), _=Depends(require_manager_or_above)):
    br = (
        db.query(BusinessRequirement)
        .filter(BusinessRequirement.id == br_id, BusinessRequirement.is_active == True)
        .first()
    )
    if not br:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Business requirement not found")
    if body.title is not None:
        br.title = body.title
    if body.location is not None:
        br.location = body.location
    if body.streams is not None:
        db.query(BRStream).filter(BRStream.br_id == br_id).update({"is_active": False})
        _bulk_insert_streams(br_id, body.streams, db)
    db.commit()
    db.refresh(br)
    return _br_full(br, db)


@router.delete("/{br_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_br(br_id: int, db: Session = Depends(get_db), _=Depends(require_manager_or_above)):
    br = (
        db.query(BusinessRequirement)
        .filter(BusinessRequirement.id == br_id, BusinessRequirement.is_active == True)
        .first()
    )
    if not br:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Business requirement not found")
    br.is_active = False
    db.commit()
