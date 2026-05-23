import asyncio
import io
import json
from datetime import datetime, timezone
from typing import Optional

import httpx
from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from auth.dependencies import get_current_user, require_manager_or_above
from config import settings
from database import get_db
from sync.models import SyncedBatch, SyncedDpiRecord, SyncedSubjectScore

from .models import ExcelBatchRegistry, TraineeStreamReference
from .schemas import RowResult, ScoresUploadResult, StreamReferenceResponse

router = APIRouter(prefix="/scores", tags=["scores-upload"])

_EXCEL_SUBJECTS = ["Java", "Python", "WebTech", "AIML", "Agile", "BizSkill"]


# ── Template ──────────────────────────────────────────────────────────

@router.get("/excel-template")
def download_template(_=Depends(get_current_user)):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Scores"

    headers = [
        "Emp Id", "Name", "Sub Batch", "DPI", "Stream",
        "Java", "Python", "WebTech", "AIML", "Agile", "BizSkill",
    ]
    hdr_fill = PatternFill(start_color="1E4D8C", end_color="1E4D8C", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    ws.append(["EMP001", "John Doe",   "A1", 3.5, "Java Development",  78.5, 65.0, 72.0, 55.0, 80.0, 70.0])
    ws.append(["EMP002", "Jane Smith", "A2", 4.2, "AI/ML Engineering", 60.0, 85.0, 70.0, 92.0, 75.0, 68.0])

    col_widths = [12, 20, 12, 8, 25, 10, 10, 12, 10, 10, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_scores_template.xlsx"},
    )


# ── Batch info / Excel-batch list ─────────────────────────────────────

@router.get("/batch-info/{batch_name}")
def batch_info(
    batch_name: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    dpi_count = (
        db.query(SyncedDpiRecord)
        .filter(SyncedDpiRecord.batch_name == batch_name)
        .count()
    )
    excel_reg = (
        db.query(ExcelBatchRegistry)
        .filter(ExcelBatchRegistry.batch_name == batch_name)
        .first()
    )
    return {
        "batch_name": batch_name,
        "dpi_count": dpi_count,
        "has_existing": dpi_count > 0,
        "excel_managed": excel_reg is not None,
        "uploaded_at": excel_reg.uploaded_at if excel_reg else None,
    }


@router.get("/excel-batches")
def list_excel_batches(db: Session = Depends(get_db), _=Depends(get_current_user)):
    rows = db.query(ExcelBatchRegistry).all()
    return [{"batch_name": r.batch_name, "uploaded_at": r.uploaded_at, "trainee_count": r.trainee_count} for r in rows]


# ── Java background backup ────────────────────────────────────────────

async def _backup_to_java(
    batch_name: str,
    dpi_payloads: list[dict],
    score_payloads: list[dict],
) -> None:
    try:
        async with httpx.AsyncClient() as client:
            # Ensure batch exists
            try:
                r = await client.get(
                    f"{settings.SPRINGBOOT_BASE_URL}/api/subjects/{batch_name}", timeout=5.0
                )
                if r.status_code == 404:
                    await client.post(
                        f"{settings.SPRINGBOOT_BASE_URL}/api/subjects",
                        json={"batchName": batch_name, "traineeCount": len(dpi_payloads), "subjects": _EXCEL_SUBJECTS},
                        timeout=5.0,
                    )
            except Exception:
                pass

            # Fire all DPI + score POSTs concurrently
            tasks = [
                client.post(f"{settings.SPRINGBOOT_BASE_URL}/api/dpi", json=p, timeout=10.0)
                for p in dpi_payloads
            ] + [
                client.post(f"{settings.SPRINGBOOT_BASE_URL}/api/scores", json=p, timeout=10.0)
                for p in score_payloads
            ]
            await asyncio.gather(*tasks, return_exceptions=True)
    except Exception:
        pass  # background backup — never raises


# ── Upload ────────────────────────────────────────────────────────────

@router.post("/upload-excel", response_model=ScoresUploadResult)
async def upload_excel(
    background_tasks: BackgroundTasks,
    batch_name: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_manager_or_above),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or corrupt Excel file")

    ws = wb.active
    row_results: list[RowResult] = []
    succeeded = 0
    failed = 0
    now = datetime.now(timezone.utc).isoformat()

    # Payloads collected for background Java backup (fired after response is sent)
    java_dpi_payloads: list[dict] = []
    java_score_payloads: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None for cell in row):
                continue

            def _get(i: int, r=row):
                return r[i] if len(r) > i else None

            emp_id     = _get(0)
            name       = _get(1)
            sub_batch  = _get(2)
            dpi_val    = _get(3)
            stream_val = _get(4)
            java       = _get(5)
            python_    = _get(6)
            webtech    = _get(7)
            aiml       = _get(8)
            agile      = _get(9)
            bizskill   = _get(10)

            if not emp_id or not name:
                row_results.append(RowResult(
                    row=row_idx, trainee_id=str(emp_id or ""),
                    status="error", detail="Emp Id and Name are required",
                ))
                failed += 1
                continue

            emp_id_str    = str(emp_id).strip()
            name_str      = str(name).strip()
            sub_batch_str = str(sub_batch).strip() if sub_batch else None

            try:
                dpi_float = float(dpi_val)
                if not (0.0 <= dpi_float <= 5.0):
                    raise ValueError
            except (TypeError, ValueError):
                row_results.append(RowResult(
                    row=row_idx, trainee_id=emp_id_str,
                    status="error", detail="DPI must be a number between 0 and 5",
                ))
                failed += 1
                continue

            subject_pairs = [
                ("Java", java), ("Python", python_), ("WebTech", webtech),
                ("AIML", aiml), ("Agile", agile), ("BizSkill", bizskill),
            ]
            subject_scores: list[tuple[str, float]] = []
            row_error: Optional[str] = None
            for subj_name, subj_val in subject_pairs:
                if subj_val is None:
                    continue
                try:
                    s = float(subj_val)
                    if not (0.0 <= s <= 100.0):
                        row_error = f"{subj_name} score must be between 0 and 100"
                        break
                    subject_scores.append((subj_name, s))
                except (TypeError, ValueError):
                    row_error = f"{subj_name} score must be a number"
                    break

            if row_error:
                row_results.append(RowResult(row=row_idx, trainee_id=emp_id_str, status="error", detail=row_error))
                failed += 1
                continue

            # ── Write directly to synced tables (Excel has priority) ──

            existing_dpi = (
                db.query(SyncedDpiRecord)
                .filter(SyncedDpiRecord.trainee_id == emp_id_str)
                .first()
            )
            if existing_dpi:
                existing_dpi.batch_name = batch_name
                existing_dpi.trainee_name = name_str
                existing_dpi.dpi = dpi_float
                existing_dpi.sub_batch = sub_batch_str
                existing_dpi.synced_at = now
            else:
                db.add(SyncedDpiRecord(
                    trainee_id=emp_id_str,
                    batch_name=batch_name,
                    trainee_name=name_str,
                    dpi=dpi_float,
                    location=None,
                    sub_batch=sub_batch_str,
                    synced_at=now,
                ))

            # Delete existing synced scores for this trainee+batch, then re-insert
            db.query(SyncedSubjectScore).filter(
                SyncedSubjectScore.trainee_id == emp_id_str,
                SyncedSubjectScore.batch_name == batch_name,
            ).delete(synchronize_session="fetch")
            for subj_name, score_val in subject_scores:
                db.add(SyncedSubjectScore(
                    external_id=f"excel-{emp_id_str}-{subj_name}",
                    batch_name=batch_name,
                    trainee_id=emp_id_str,
                    trainee_name=name_str,
                    subject_name=subj_name,
                    subject_id=None,
                    exam_name="Excel Upload",
                    score=score_val,
                    synced_at=now,
                ))

            # ── Collect Java backup payloads (sent after response via background task) ──
            java_dpi_payloads.append({
                "traineeId": emp_id_str, "batchName": batch_name,
                "traineeName": name_str, "dpi": dpi_float,
                "location": None, "subBatch": sub_batch_str,
            })
            for subj_name, score_val in subject_scores:
                java_score_payloads.append({
                    "batchName": batch_name, "traineeId": emp_id_str,
                    "traineeName": name_str, "subjectName": subj_name,
                    "subjectId": None, "examName": "Excel Upload", "score": score_val,
                })

            # ── Stream reference ──
            if stream_val:
                stream_str = str(stream_val).strip()
                if stream_str:
                    ref = (
                        db.query(TraineeStreamReference)
                        .filter(TraineeStreamReference.trainee_id == emp_id_str)
                        .first()
                    )
                    if ref:
                        ref.batch_name = batch_name
                        ref.stream_name = stream_str
                        ref.updated_at = now
                    else:
                        db.add(TraineeStreamReference(
                            trainee_id=emp_id_str,
                            batch_name=batch_name,
                            stream_name=stream_str,
                            updated_at=now,
                        ))

            row_results.append(RowResult(row=row_idx, trainee_id=emp_id_str, status="ok"))
            succeeded += 1

    # Upsert synced_batches entry so the batch appears everywhere
    existing_batch = (
        db.query(SyncedBatch).filter(SyncedBatch.batch_name == batch_name).first()
    )
    if existing_batch:
        existing_batch.subjects_json = json.dumps(_EXCEL_SUBJECTS)
        if succeeded > existing_batch.trainee_count:
            existing_batch.trainee_count = succeeded
        existing_batch.synced_at = now
    else:
        db.add(SyncedBatch(
            batch_name=batch_name,
            subjects_json=json.dumps(_EXCEL_SUBJECTS),
            trainee_count=succeeded,
            synced_at=now,
        ))

    # Register / update Excel batch registry
    reg = (
        db.query(ExcelBatchRegistry)
        .filter(ExcelBatchRegistry.batch_name == batch_name)
        .first()
    )
    if reg:
        reg.uploaded_at = now
        reg.trainee_count = succeeded
    else:
        db.add(ExcelBatchRegistry(
            batch_name=batch_name,
            uploaded_at=now,
            trainee_count=succeeded,
        ))

    db.commit()

    # Schedule Java backup — runs after response is sent, all rows concurrently
    background_tasks.add_task(_backup_to_java, batch_name, java_dpi_payloads, java_score_payloads)

    return ScoresUploadResult(
        rows_processed=succeeded + failed,
        rows_succeeded=succeeded,
        rows_failed=failed,
        row_results=row_results,
        sync_triggered=False,  # direct write — no sync step needed
    )


# ── Stream references ─────────────────────────────────────────────────

@router.get("/stream-references", response_model=list[StreamReferenceResponse])
def list_stream_references(
    batch_name: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(TraineeStreamReference)
    if batch_name:
        q = q.filter(TraineeStreamReference.batch_name == batch_name)
    return q.all()
